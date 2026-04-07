
import io
import json
import re

import openpyxl
import pandas as pd
import streamlit as st

st.set_page_config(
    page_title="Pliego Tarifario SAESA",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded",
)

# =============================================================================
# ESTILO VISUAL CORREGIDO
# =============================================================================
st.markdown("""
<style>
:root{
    --bg: #f3f6fb;
    --card: #ffffff;
    --card-2: #f8fafc;
    --text: #111827;
    --muted: #6b7280;
    --primary: #2563eb;
    --primary-2: #1d4ed8;
    --accent: #7c3aed;
    --border: #dbe3ef;
    --sidebar: #0f172a;
    --sidebar-2: #111827;
    --success: #16a34a;
}

html, body, [data-testid="stAppViewContainer"] {
    background: var(--bg) !important;
    color: var(--text) !important;
}

.stApp {
    background: linear-gradient(180deg, #eef3f9 0%, #f8fbff 100%) !important;
}

[data-testid="stHeader"]{
    background: rgba(255,255,255,0.0) !important;
}

[data-testid="stSidebar"] {
    background: linear-gradient(180deg, var(--sidebar) 0%, var(--sidebar-2) 100%) !important;
    border-right: 1px solid rgba(255,255,255,.08);
}

[data-testid="stSidebar"] * {
    color: #f8fafc !important;
}

[data-testid="stFileUploader"] section {
    background: rgba(255,255,255,.04) !important;
    border: 1px solid rgba(255,255,255,.10) !important;
}

.block-container {
    padding-top: 1.4rem;
    padding-bottom: 2rem;
    max-width: 1500px;
}

.main-title {
    font-size: 2.3rem;
    font-weight: 800;
    color: var(--text);
    margin-bottom: .15rem;
    letter-spacing: -0.02em;
}
.main-subtitle {
    color: var(--muted);
    font-size: 1rem;
    margin-bottom: 1.2rem;
}
.panel {
    background: var(--card);
    border: 1px solid var(--border);
    border-radius: 16px;
    padding: 1rem 1.15rem;
    box-shadow: 0 8px 24px rgba(15, 23, 42, .05);
}
.metric-card {
    background: var(--card);
    border: 1px solid var(--border);
    border-radius: 16px;
    padding: 1rem 1.1rem;
    box-shadow: 0 6px 18px rgba(15, 23, 42, .05);
    min-height: 110px;
}
.metric-label {
    color: var(--muted);
    font-size: .85rem;
    margin-bottom: .35rem;
    font-weight: 600;
}
.metric-value {
    color: var(--primary-2);
    font-size: 2rem;
    font-weight: 800;
    line-height: 1.05;
}
.metric-sub {
    color: var(--muted);
    font-size: .82rem;
    margin-top: .35rem;
}
.period-badge {
    display: inline-block;
    background: #e0ecff;
    color: #1e40af;
    border: 1px solid #bfdbfe;
    border-radius: 999px;
    padding: 4px 12px;
    margin: 4px 6px 0 0;
    font-size: .82rem;
    font-weight: 700;
}
.tariff-header {
    background: linear-gradient(90deg, #5b21b6 0%, #7c3aed 100%);
    color: white;
    padding: 10px 14px;
    border-radius: 12px;
    font-weight: 800;
    margin: 18px 0 8px 0;
    border: 1px solid rgba(255,255,255,.15);
}
.section-title {
    color: var(--text);
    font-size: 1.5rem;
    font-weight: 800;
    margin-bottom: .8rem;
}
.stTabs [data-baseweb="tab-list"] {
    gap: .35rem;
    border-bottom: 1px solid var(--border);
}
.stTabs [data-baseweb="tab"] {
    border-radius: 12px 12px 0 0;
    background: #eef2ff;
    color: #334155 !important;
    border: 1px solid var(--border);
    border-bottom: none !important;
    padding: .6rem 1rem;
    font-weight: 700;
}
.stTabs [aria-selected="true"] {
    background: white !important;
    color: var(--primary-2) !important;
}
div[data-baseweb="select"] > div,
div[data-baseweb="input"] > div,
.stTextInput input {
    background: #ffffff !important;
    color: var(--text) !important;
    border: 1px solid var(--border) !important;
    border-radius: 12px !important;
}
.stMultiSelect [data-baseweb="tag"]{
    background: #e0ecff !important;
    border-radius: 999px !important;
    border: 1px solid #bfdbfe !important;
}
.stRadio label, .stSelectbox label, .stMultiSelect label, .stTextInput label {
    color: var(--text) !important;
    font-weight: 700 !important;
}
.stButton>button, .stDownloadButton>button {
    border-radius: 12px !important;
    border: 1px solid transparent !important;
    font-weight: 700 !important;
}
.stDownloadButton>button {
    background: linear-gradient(90deg, var(--primary), var(--accent)) !important;
    color: white !important;
}
.stAlert {
    border-radius: 14px !important;
}
[data-testid="stDataFrame"] {
    background: white !important;
    border: 1px solid var(--border);
    border-radius: 14px;
}
.small-note {
    color: var(--muted);
    font-size: .86rem;
}
</style>
""", unsafe_allow_html=True)

MESES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4,
    "mayo": 5, "junio": 6, "julio": 7, "agosto": 8,
    "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12,
}
TARIFF_PATTERN = re.compile(r"^\s*Tarifa\s+((BT|AT|TRBT|TRAT).*)", re.IGNORECASE)

# =============================================================================
# HELPERS
# =============================================================================
def periodo_sort_key(periodo: str):
    if not isinstance(periodo, str):
        return (9999, 99)
    m = re.search(r"(enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|octubre|noviembre|diciembre)\s+(\d{4})", periodo.lower())
    if not m:
        return (9999, 99)
    return (int(m.group(2)), MESES[m.group(1)])


def clean_text(val):
    if val is None:
        return None
    return str(val).replace("\n", " ").replace("\r", " ").strip()


def to_num(v):
    if isinstance(v, (int, float)):
        return float(v)
    return None


def choose_sheet(workbook):
    """Prioriza 'Pub. Zonal'. Si no existe, usa la primera hoja visible."""
    for preferred in ["Pub. Zonal", "PUB. ZONAL", "Pub.Zonal"]:
        if preferred in workbook.sheetnames:
            return workbook[preferred], preferred

    for ws in workbook.worksheets:
        if ws.sheet_state == "visible":
            return ws, ws.title

    return workbook.active, workbook.active.title


def find_period(ws):
    for r in range(1, min(ws.max_row, 20) + 1):
        for c in range(1, min(ws.max_column, 8) + 1):
            txt = clean_text(ws.cell(r, c).value)
            if not txt:
                continue
            low = txt.lower()
            m = re.search(r"(enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|octubre|noviembre|diciembre)\s+de\s+(\d{4})", low)
            if "tarifas de suministro" in low and m:
                return f"{m.group(1)} {m.group(2)}"
    return "desconocido"


def find_localidades_row(ws):
    for r in range(1, min(ws.max_row, 20) + 1):
        matches = 0
        for c in range(4, min(ws.max_column, 40) + 1):
            txt = clean_text(ws.cell(r, c).value)
            if txt and (" - Aéreo" in txt or " - Subterráneo" in txt):
                matches += 1
        if matches >= 2:
            return r
    return 5


def extract_localidades(ws, row_idx):
    localidades = {}
    col = 4
    while col <= ws.max_column:
        txt = clean_text(ws.cell(row_idx, col).value)
        if txt and (" - Aéreo" in txt or " - Subterráneo" in txt):
            localidades[col] = txt
            # En estos pliegos el valor neto está en esta col y c/iva en la siguiente
            col += 2
        else:
            col += 1
    return localidades


def find_data_start(ws, comunas_row):
    for r in range(comunas_row + 1, min(comunas_row + 8, ws.max_row) + 1):
        vals = [clean_text(ws.cell(r, c).value) for c in range(3, min(ws.max_column, 12) + 1)]
        if any(v and "$ NETO" in v.upper() for v in vals):
            return r + 1
    return comunas_row + 3


def find_tariff_headers(ws, data_start):
    headers = []
    for r in range(data_start, ws.max_row + 1):
        txt = clean_text(ws.cell(r, 2).value)
        c3 = ws.cell(r, 3).value
        if txt and c3 is None and TARIFF_PATTERN.match(txt):
            headers.append((r, txt))
    return headers


def find_injection_start(ws):
    for r in range(1, ws.max_row + 1):
        txt = clean_text(ws.cell(r, 2).value)
        if txt and txt.lower().startswith("precios para valorización de inyecciones de energía"):
            return r
    return None


def split_localidad(localidad):
    parts = localidad.rsplit(" - ", 1)
    if len(parts) == 2:
        return parts[0], parts[1]
    return localidad, ""


# =============================================================================
# PARSER
# =============================================================================
@st.cache_data(show_spinner=False)
def procesar_excel(file_bytes: bytes, file_name: str):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws, sheet_used = choose_sheet(wb)

    periodo = find_period(ws)
    comunas_row = find_localidades_row(ws)
    localidades = extract_localidades(ws, comunas_row)
    data_start = find_data_start(ws, comunas_row)
    injection_start = find_injection_start(ws)
    tariff_headers = find_tariff_headers(ws, data_start)

    if not localidades:
        raise ValueError(f"No se encontraron localidades válidas en la hoja '{sheet_used}'")
    if not tariff_headers:
        raise ValueError(f"No se encontraron encabezados de tarifas en la hoja '{sheet_used}'")

    rows = []

    # Construye bloques por tarifa
    for i, (header_row, tarifa) in enumerate(tariff_headers):
        next_header_row = tariff_headers[i + 1][0] if i + 1 < len(tariff_headers) else (injection_start or ws.max_row + 1)
        end_row = next_header_row - 1

        for r in range(header_row + 1, end_row + 1):
            concepto = clean_text(ws.cell(r, 2).value)
            unidad = clean_text(ws.cell(r, 3).value)

            if not concepto:
                continue
            # Evitar notas/pies dentro de bloques
            if unidad is None:
                continue

            for loc_col, localidad in localidades.items():
                neto = to_num(ws.cell(r, loc_col).value)
                civa = to_num(ws.cell(r, loc_col + 1).value)
                comuna, tipo_suministro = split_localidad(localidad)

                rows.append({
                    "archivo": file_name,
                    "hoja_origen": sheet_used,
                    "periodo": periodo,
                    "tarifa": tarifa,
                    "concepto": concepto,
                    "unidad": unidad,
                    "localidad": localidad,
                    "comuna": comuna,
                    "tipo_suministro": tipo_suministro,
                    "valor_neto": round(neto, 6) if neto is not None else None,
                    "valor_civa": round(civa, 6) if civa is not None else None,
                })

    df_tarifas = pd.DataFrame(rows)

    # Inyecciones
    inj_rows = []
    if injection_start:
        inj_localidades_row = None
        for r in range(injection_start, min(injection_start + 8, ws.max_row) + 1):
            if any(
                clean_text(ws.cell(r, c).value) and (" - Aéreo" in clean_text(ws.cell(r, c).value) or " - Subterráneo" in clean_text(ws.cell(r, c).value))
                for c in range(4, min(ws.max_column, 40) + 1)
            ):
                inj_localidades_row = r
                break

        inj_localidades = extract_localidades(ws, inj_localidades_row) if inj_localidades_row else localidades

        for r in range((inj_localidades_row or injection_start) + 1, ws.max_row + 1):
            concepto = clean_text(ws.cell(r, 2).value)
            unidad = clean_text(ws.cell(r, 3).value)

            if not concepto:
                continue

            low = concepto.lower()
            if "los valores indicados son netos" in low or "sociedad austral de electricidad" in low:
                break

            if unidad is None:
                continue

            for loc_col, localidad in inj_localidades.items():
                neto = to_num(ws.cell(r, loc_col).value)
                if neto is None:
                    continue
                comuna, tipo_suministro = split_localidad(localidad)
                inj_rows.append({
                    "archivo": file_name,
                    "hoja_origen": sheet_used,
                    "periodo": periodo,
                    "concepto": concepto,
                    "unidad": unidad,
                    "localidad": localidad,
                    "comuna": comuna,
                    "tipo_suministro": tipo_suministro,
                    "valor_neto": round(neto, 6),
                })

    df_iny = pd.DataFrame(inj_rows)

    # Catálogo útil para control
    resumen = {
        "archivo": file_name,
        "hoja_usada": sheet_used,
        "periodo": periodo,
        "tarifas_detectadas": sorted(df_tarifas["tarifa"].dropna().unique().tolist()) if not df_tarifas.empty else [],
        "n_tarifas": int(df_tarifas["tarifa"].nunique()) if not df_tarifas.empty else 0,
        "n_localidades": int(df_tarifas["localidad"].nunique()) if not df_tarifas.empty else 0,
        "n_registros": int(len(df_tarifas)),
    }
    return df_tarifas, df_iny, resumen


# =============================================================================
# UI SIDEBAR
# =============================================================================
with st.sidebar:
    st.markdown("## ⚡ SAESA Tarifas")
    st.caption("Parser corregido para hoja Pub. Zonal")
    st.markdown("---")
    uploaded_files = st.file_uploader(
        "Sube uno o más pliegos tarifarios Excel",
        type=["xlsx"],
        accept_multiple_files=True,
        help="La app prioriza la hoja 'Pub. Zonal' y extrae tarifas BT*, AT*, TRBT*, TRAT* e inyecciones.",
    )
    st.markdown("---")
    st.markdown('<div class="small-note">El parser usa <b>Pub. Zonal</b> como hoja principal. Si no existe, usa la primera hoja visible.</div>', unsafe_allow_html=True)

# =============================================================================
# TITULAR
# =============================================================================
st.markdown('<div class="main-title">⚡ Pliego Tarifario SAESA</div>', unsafe_allow_html=True)
st.markdown(
    '<div class="main-subtitle">Versión corregida para leer la hoja <b>Pub. Zonal</b>, capturar tarifas <b>BT*</b>, <b>AT*</b>, <b>TRBT*</b>, <b>TRAT*</b> y mostrar la plataforma con un fondo limpio y legible.</div>',
    unsafe_allow_html=True
)

if not uploaded_files:
    st.info("Carga uno o más archivos Excel desde el panel izquierdo para comenzar.")
    st.stop()

# =============================================================================
# PROCESAMIENTO
# =============================================================================
all_tarifas, all_iny, all_resumen, errors = [], [], [], []

progress = st.progress(0, text="Procesando archivos...")
for i, f in enumerate(uploaded_files):
    try:
        file_bytes = f.read()
        df_t, df_i, resumen = procesar_excel(file_bytes, f.name)
        all_tarifas.append(df_t)
        if not df_i.empty:
            all_iny.append(df_i)
        all_resumen.append(resumen)
        progress.progress((i + 1) / len(uploaded_files), text=f"✅ {f.name} — {resumen['periodo']} — hoja {resumen['hoja_usada']}")
    except Exception as e:
        errors.append(f"{f.name}: {e}")
        progress.progress((i + 1) / len(uploaded_files), text=f"⚠️ {f.name}")

progress.empty()

for err in errors:
    st.error(err)

if not all_tarifas:
    st.error("No se pudo procesar ningún archivo.")
    st.stop()

df_all = pd.concat(all_tarifas, ignore_index=True)
df_iny_all = pd.concat(all_iny, ignore_index=True) if all_iny else pd.DataFrame()
df_resumen = pd.DataFrame(all_resumen)

periodos_ordenados = sorted(df_all["periodo"].dropna().unique().tolist(), key=periodo_sort_key)
df_all["periodo"] = pd.Categorical(df_all["periodo"], categories=periodos_ordenados, ordered=True)
df_all = df_all.sort_values(["periodo", "tarifa", "localidad", "concepto"]).reset_index(drop=True)

localidades_list = sorted(df_all["localidad"].dropna().unique().tolist())
comunas_list = sorted(df_all["comuna"].dropna().unique().tolist())
tarifas_list = sorted(df_all["tarifa"].dropna().unique().tolist())

# =============================================================================
# MÉTRICAS
# =============================================================================
m1, m2, m3, m4, m5 = st.columns(5)
metrics = [
    ("Períodos", len(periodos_ordenados), "Archivos tarifarios leídos"),
    ("Localidades", df_all["localidad"].nunique(), "Pub. Zonal"),
    ("Tarifas", df_all["tarifa"].nunique(), "BT, AT, TRBT, TRAT"),
    ("Conceptos", df_all["concepto"].nunique(), "Cargos y componentes"),
    ("Registros", f"{len(df_all):,}", "Base normalizada"),
]
for col, (label, value, sub) in zip([m1, m2, m3, m4, m5], metrics):
    col.markdown(
        f"""
        <div class="metric-card">
            <div class="metric-label">{label}</div>
            <div class="metric-value">{value}</div>
            <div class="metric-sub">{sub}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

periodos_badges = "".join([f'<span class="period-badge">{p}</span>' for p in periodos_ordenados])
st.markdown(f"<div class='panel'><b>Períodos cargados:</b><br>{periodos_badges}</div>", unsafe_allow_html=True)

# Control parser
with st.expander("Ver control de lectura por archivo"):
    st.dataframe(df_resumen, use_container_width=True, hide_index=True)
    if not df_resumen.empty:
        st.write("Tarifas detectadas por archivo:")
        for _, row in df_resumen.iterrows():
            st.markdown(f"**{row['archivo']}** · hoja usada: `{row['hoja_usada']}` · tarifas: `{', '.join(row['tarifas_detectadas'])}`")

st.markdown("<br>", unsafe_allow_html=True)

# =============================================================================
# TABS
# =============================================================================
tabs = st.tabs([
    "🔍 Consulta por Localidad",
    "📊 Comparar Localidades",
    "📈 Evolución Temporal",
    "🗄️ Base de Datos",
    "🔋 Inyecciones",
])

# -----------------------------------------------------------------------------
# TAB 1
# -----------------------------------------------------------------------------
with tabs[0]:
    st.markdown('<div class="section-title">Consulta de tarifas por localidad</div>', unsafe_allow_html=True)
    c1, c2, c3, c4 = st.columns([2, 2, 1, 1])
    with c1:
        sel_loc = st.selectbox("Localidad", localidades_list)
    with c2:
        sel_tar = st.selectbox("Tarifa", ["Todas"] + tarifas_list)
    with c3:
        sel_per = st.selectbox("Período", ["Último"] + periodos_ordenados)
    with c4:
        pt1 = st.radio("Precio", ["Neto", "C/IVA"], horizontal=True)

    pcol = "valor_neto" if pt1 == "Neto" else "valor_civa"
    per_filt = periodos_ordenados[-1] if sel_per == "Último" else sel_per

    df_loc = df_all[(df_all["localidad"] == sel_loc) & (df_all["periodo"] == per_filt)].copy()
    if sel_tar != "Todas":
        df_loc = df_loc[df_loc["tarifa"] == sel_tar]

    if df_loc.empty:
        st.warning("No hay datos para la selección.")
    else:
        for tarifa_name, grp in df_loc.groupby("tarifa", sort=False):
            st.markdown(f'<div class="tariff-header">📋 {tarifa_name}</div>', unsafe_allow_html=True)
            show = grp[["concepto", "unidad", pcol]].copy()
            show.columns = ["Concepto", "Unidad", f"$ {pt1}"]
            st.dataframe(
                show.style.format({f"$ {pt1}": lambda x: f"{x:,.3f}" if pd.notna(x) else "—"}),
                use_container_width=True,
                hide_index=True,
            )

# -----------------------------------------------------------------------------
# TAB 2
# -----------------------------------------------------------------------------
with tabs[1]:
    st.markdown('<div class="section-title">Comparación entre localidades</div>', unsafe_allow_html=True)
    c1, c2 = st.columns([3, 2])
    with c1:
        locs_sel = st.multiselect("Localidades (máx 6)", localidades_list, default=localidades_list[:3], max_selections=6)
    with c2:
        tar_comp = st.selectbox("Tarifa", tarifas_list)
        per_comp = st.selectbox("Período", ["Último"] + periodos_ordenados)
        pc2 = st.radio("Precio", ["Neto", "C/IVA"], horizontal=True, key="cmp_precio")

    pcol2 = "valor_neto" if pc2 == "Neto" else "valor_civa"
    per_c2 = periodos_ordenados[-1] if per_comp == "Último" else per_comp

    if not locs_sel:
        st.info("Selecciona al menos una localidad.")
    else:
        df_comp = df_all[
            (df_all["tarifa"] == tar_comp) &
            (df_all["localidad"].isin(locs_sel)) &
            (df_all["periodo"] == per_c2)
        ].copy()

        if df_comp.empty:
            st.warning("Sin datos para la selección.")
        else:
            pivot = df_comp.pivot_table(
                index=["concepto", "unidad"],
                columns="localidad",
                values=pcol2,
                aggfunc="first",
            ).reset_index()
            pivot.columns.name = None
            num_cols = [c for c in pivot.columns if c not in ["concepto", "unidad"]]
            st.dataframe(
                pivot.style.format({c: lambda x: f"{x:,.3f}" if pd.notna(x) else "—" for c in num_cols}),
                use_container_width=True,
                hide_index=True,
            )

# -----------------------------------------------------------------------------
# TAB 3
# -----------------------------------------------------------------------------
with tabs[2]:
    st.markdown('<div class="section-title">Evolución temporal de tarifas</div>', unsafe_allow_html=True)

    if len(periodos_ordenados) < 2:
        st.info("Carga al menos dos períodos para ver evolución.")
    else:
        ec1, ec2, ec3, ec4 = st.columns([2, 2, 2, 1])
        with ec1:
            evo_loc = st.selectbox("Localidad", localidades_list, key="evo_loc")
        with ec2:
            evo_tar = st.selectbox("Tarifa", tarifas_list, key="evo_tar")
        with ec3:
            conceptos_disp = sorted(df_all[df_all["tarifa"] == evo_tar]["concepto"].dropna().unique().tolist())
            evo_conc = st.selectbox("Concepto", conceptos_disp, key="evo_conc")
        with ec4:
            evo_p = st.radio("Precio", ["Neto", "C/IVA"], horizontal=True, key="evo_precio")

        evo_col = "valor_neto" if evo_p == "Neto" else "valor_civa"
        df_evo = df_all[
            (df_all["localidad"] == evo_loc) &
            (df_all["tarifa"] == evo_tar) &
            (df_all["concepto"] == evo_conc)
        ][["periodo", evo_col]].dropna().copy()

        subtabs = st.tabs(["📈 Gráfico", "📋 Tabla mes a mes", "📉 Variación %", "🚨 Alertas"])

        with subtabs[0]:
            if df_evo.empty:
                st.warning("Sin datos para la combinación seleccionada.")
            else:
                unidad = df_all[
                    (df_all["tarifa"] == evo_tar) &
                    (df_all["concepto"] == evo_conc)
                ]["unidad"].dropna().iloc[0]

                labels = df_evo["periodo"].astype(str).tolist()
                values = df_evo[evo_col].tolist()
                chart_data = json.dumps({"labels": labels, "values": values})

                chart_html = f"""
                <!DOCTYPE html>
                <html>
                <head>
                <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
                </head>
                <body style="margin:0;background:#ffffff;font-family:Arial,sans-serif;">
                    <canvas id="chart" height="120"></canvas>
                    <script>
                    const d = {chart_data};
                    new Chart(document.getElementById('chart'), {{
                        type: 'line',
                        data: {{
                            labels: d.labels,
                            datasets: [{{
                                label: {json.dumps(f"$ {evo_p} ({unidad})")},
                                data: d.values,
                                borderColor: '#2563eb',
                                backgroundColor: 'rgba(37,99,235,0.10)',
                                borderWidth: 3,
                                pointRadius: 5,
                                pointHoverRadius: 7,
                                tension: 0.25,
                                fill: true
                            }}]
                        }},
                        options: {{
                            responsive: true,
                            plugins: {{
                                legend: {{ display: true }},
                                title: {{
                                    display: true,
                                    text: {json.dumps(f"{evo_conc} — {evo_loc}")},
                                    color: '#111827',
                                    font: {{ size: 16, weight: 'bold' }}
                                }}
                            }},
                            scales: {{
                                x: {{
                                    ticks: {{ color: '#475569' }},
                                    grid: {{ color: '#e5e7eb' }}
                                }},
                                y: {{
                                    ticks: {{
                                        color: '#475569',
                                        callback: function(value) {{
                                            return '$' + value.toLocaleString('es-CL');
                                        }}
                                    }},
                                    grid: {{ color: '#e5e7eb' }}
                                }}
                            }}
                        }}
                    }});
                    </script>
                </body>
                </html>
                """
                st.components.v1.html(chart_html, height=420)
                st.dataframe(
                    df_evo.rename(columns={evo_col: f"$ {evo_p}", "periodo": "Período"}).style.format({f"$ {evo_p}": "{:,.3f}"}),
                    use_container_width=True,
                    hide_index=True,
                )

        with subtabs[1]:
            df_pt = df_all[
                (df_all["localidad"] == evo_loc) &
                (df_all["tarifa"] == evo_tar)
            ].pivot_table(
                index=["concepto", "unidad"],
                columns="periodo",
                values=evo_col,
                aggfunc="first",
            ).reset_index()
            df_pt.columns.name = None
            num_cols = [c for c in df_pt.columns if c not in ["concepto", "unidad"]]
            st.dataframe(
                df_pt.style.format({c: lambda x: f"{x:,.3f}" if pd.notna(x) else "—" for c in num_cols}),
                use_container_width=True,
                hide_index=True,
            )

        with subtabs[2]:
            cva, cvb = st.columns(2)
            with cva:
                per_base = st.selectbox("Período base", periodos_ordenados[:-1], key="var_base")
            with cvb:
                cand = [p for p in periodos_ordenados if periodo_sort_key(p) > periodo_sort_key(per_base)]
                per_cmp = st.selectbox("Período a comparar", cand, key="var_cmp")

            def get_vals(per):
                return df_all[
                    (df_all["localidad"] == evo_loc) &
                    (df_all["tarifa"] == evo_tar) &
                    (df_all["periodo"] == per)
                ][["concepto", "unidad", evo_col]]

            df_var = get_vals(per_base).rename(columns={evo_col: "base"}).merge(
                get_vals(per_cmp).rename(columns={evo_col: "nuevo"}),
                on=["concepto", "unidad"],
                how="outer"
            )
            df_var["variacion_%"] = ((df_var["nuevo"] - df_var["base"]) / df_var["base"] * 100).round(2)
            df_var = df_var.rename(columns={
                "concepto": "Concepto",
                "unidad": "Unidad",
                "base": f"$ {per_base}",
                "nuevo": f"$ {per_cmp}",
                "variacion_%": "Variación %",
            })
            st.dataframe(
                df_var.style.format({
                    f"$ {per_base}": lambda x: f"{x:,.3f}" if pd.notna(x) else "—",
                    f"$ {per_cmp}": lambda x: f"{x:,.3f}" if pd.notna(x) else "—",
                    "Variación %": lambda x: f"{x:+.2f}%" if pd.notna(x) else "—",
                }),
                use_container_width=True,
                hide_index=True,
            )

        with subtabs[3]:
            umbral = st.slider("Umbral mínimo de variación (%)", 0.0, 30.0, 1.0, 0.5)
            per_ant = periodos_ordenados[-2]
            per_ult = periodos_ordenados[-1]

            df_alert = df_all[df_all["periodo"] == per_ant][["localidad", "tarifa", "concepto", "unidad", evo_col]].rename(columns={evo_col: "ant"}).merge(
                df_all[df_all["periodo"] == per_ult][["localidad", "tarifa", "concepto", "unidad", evo_col]].rename(columns={evo_col: "ult"}),
                on=["localidad", "tarifa", "concepto", "unidad"],
                how="outer"
            )
            df_alert["variacion_%"] = ((df_alert["ult"] - df_alert["ant"]) / df_alert["ant"] * 100).round(2)
            df_alert = df_alert[df_alert["variacion_%"].abs() >= umbral].dropna(subset=["variacion_%"]).copy()

            f1, f2 = st.columns(2)
            with f1:
                floc = st.selectbox("Filtrar localidad", ["Todas"] + comunas_list, key="alerta_loc")
            with f2:
                ftar = st.selectbox("Filtrar tarifa", ["Todas"] + tarifas_list, key="alerta_tar")

            if floc != "Todas":
                df_alert = df_alert[df_alert["localidad"].str.startswith(floc, na=False)]
            if ftar != "Todas":
                df_alert = df_alert[df_alert["tarifa"] == ftar]

            df_alert = df_alert.rename(columns={
                "localidad": "Localidad",
                "tarifa": "Tarifa",
                "concepto": "Concepto",
                "unidad": "Unidad",
                "ant": f"$ {per_ant}",
                "ult": f"$ {per_ult}",
                "variacion_%": "Variación %",
            })

            st.caption(f"{len(df_alert)} registros con variación absoluta mayor o igual a {umbral}%")
            st.dataframe(
                df_alert.style.format({
                    f"$ {per_ant}": lambda x: f"{x:,.3f}" if pd.notna(x) else "—",
                    f"$ {per_ult}": lambda x: f"{x:,.3f}" if pd.notna(x) else "—",
                    "Variación %": lambda x: f"{x:+.2f}%" if pd.notna(x) else "—",
                }),
                use_container_width=True,
                hide_index=True,
                height=460,
            )

# -----------------------------------------------------------------------------
# TAB 4
# -----------------------------------------------------------------------------
with tabs[3]:
    st.markdown('<div class="section-title">Base de datos normalizada completa</div>', unsafe_allow_html=True)
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        f_per = st.multiselect("Período", periodos_ordenados)
    with c2:
        f_com = st.multiselect("Comuna", comunas_list)
    with c3:
        f_tar = st.multiselect("Tarifa", tarifas_list)
    with c4:
        f_txt = st.text_input("Buscar concepto", placeholder="ej: energía")

    df_full = df_all.copy()
    if f_per:
        df_full = df_full[df_full["periodo"].isin(f_per)]
    if f_com:
        df_full = df_full[df_full["comuna"].isin(f_com)]
    if f_tar:
        df_full = df_full[df_full["tarifa"].isin(f_tar)]
    if f_txt:
        df_full = df_full[df_full["concepto"].str.contains(f_txt, case=False, na=False)]

    st.caption(f"Mostrando {len(df_full):,} registros")
    st.dataframe(
        df_full[[
            "archivo", "hoja_origen", "periodo", "tarifa", "concepto", "unidad",
            "localidad", "comuna", "tipo_suministro", "valor_neto", "valor_civa"
        ]].style.format({
            "valor_neto": lambda x: f"{x:,.3f}" if pd.notna(x) else "—",
            "valor_civa": lambda x: f"{x:,.3f}" if pd.notna(x) else "—",
        }),
        use_container_width=True,
        hide_index=True,
        height=520,
    )

    st.download_button(
        "⬇️ Descargar CSV filtrado",
        data=df_full.to_csv(index=False).encode("utf-8-sig"),
        file_name="tarifas_saesa_filtrado.csv",
        mime="text/csv",
    )

# -----------------------------------------------------------------------------
# TAB 5
# -----------------------------------------------------------------------------
with tabs[4]:
    st.markdown('<div class="section-title">Precios para valorización de inyecciones</div>', unsafe_allow_html=True)

    if df_iny_all.empty:
        st.warning("No se encontraron datos de inyecciones.")
    else:
        c1, c2 = st.columns(2)
        with c1:
            per_iny = st.selectbox("Período", ["Último"] + periodos_ordenados)
        with c2:
            locs_iny = sorted(df_iny_all["localidad"].dropna().unique().tolist())
            f_loc_iny = st.multiselect("Filtrar localidad", locs_iny)

        per_i = periodos_ordenados[-1] if per_iny == "Último" else per_iny
        df_iny_f = df_iny_all[df_iny_all["periodo"] == per_i].copy()
        if f_loc_iny:
            df_iny_f = df_iny_f[df_iny_f["localidad"].isin(f_loc_iny)]

        if df_iny_f.empty:
            st.warning("Sin datos.")
        else:
            pivot_iny = df_iny_f.pivot_table(
                index=["concepto", "unidad"],
                columns="localidad",
                values="valor_neto",
                aggfunc="first"
            ).reset_index()
            pivot_iny.columns.name = None
            num_cols = [c for c in pivot_iny.columns if c not in ["concepto", "unidad"]]
            st.dataframe(
                pivot_iny.style.format({c: lambda x: f"{x:,.3f}" if pd.notna(x) else "—" for c in num_cols}),
                use_container_width=True,
                hide_index=True,
            )

# =============================================================================
# DESCARGA BD COMPLETA
# =============================================================================
st.markdown("<br>", unsafe_allow_html=True)
st.download_button(
    "⬇️ Descargar base completa normalizada",
    data=df_all.to_csv(index=False).encode("utf-8-sig"),
    file_name="tarifas_saesa_base_completa.csv",
    mime="text/csv",
)
